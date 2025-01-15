from flask import Flask, render_template, request, session, send_file, jsonify
import pandas as pd
from io import BytesIO
import json

# Fonction pour chercher le prix unitaire et l'unité
def find_prix_unitaire(description):
    with open('data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Parcours uniquement les sous-détails
    for lot in data['lots']:
        for sous_detail in lot['sous_details']:
            if sous_detail['description'].strip().lower() == description.strip().lower():
                return sous_detail['prix_unitaire'], sous_detail['unite']
    
    return 0, "U"

app = Flask(__name__)
app.secret_key = 'dynamic_lots_secret_key'

@app.route('/test-prix-unitaire')
def test_prix_unitaire():
    lot_name = "Gros Œuvre"
    description = "Plancher"
    prix_unitaire, unite = find_prix_unitaire(lot_name, description)
    return f"Prix unitaire: {prix_unitaire}, Unité: {unite}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get-prix-unitaire', methods=['POST'])
def get_prix_unitaire():
    data = request.json
    lot_name = data.get('lot_name')
    description = data.get('description')

    prix_unitaire, unite = find_prix_unitaire(lot_name, description)
    if prix_unitaire is None or unite is None:
        return jsonify({'error': 'Lot ou sous-détail introuvable'}), 404

    return jsonify({
        'prix_unitaire': prix_unitaire,
        'unite': unite
    })

@app.route('/add-lot', methods=['POST'])
def add_lot():
    new_lot = request.json  # Attendre un objet JSON avec 'nom' et 'sous_details'

    # Charger et mettre à jour le fichier JSON
    with open('data.json', 'r') as f:
        db = json.load(f)

    db['lots'].append(new_lot)

    with open('data.json', 'w') as f:
        json.dump(db, f, indent=4)

    return jsonify({'message': 'Lot ajouté avec succès !'})

@app.route('/data-entry', methods=['GET', 'POST'])
def data_entry():
    if request.method == 'POST':
        # Récupération des données pour les lots principaux
        lots = request.form.getlist('lot_name[]')
        descriptions = request.form.getlist('description[]')
        quantities = request.form.getlist('quantity[]')
        unit = request.form.getlist('unit[]')

        # Récupération des données pour les sous-détails
        sub_descriptions = request.form.getlist('sub_description[]')
        sub_quantities = request.form.getlist('sub_quantity[]')
        sub_units = request.form.getlist('sub_unit[]')
        sub_unit_prices = request.form.getlist('sub_unit_price[]')

        # Construire les données des lots principaux
        all_lots = []
        total_global = 0

        for i in range(len(lots)):
            try:
                lot_name = lots[i]
                description = descriptions[i]
                quantity = float(quantities[i]) if quantities[i] else 0

                prix_unitaire, unit = find_prix_unitaire(description)
                    print(f"Description: {description}, Prix Unitaire: {prix_unitaire}, Unité: {unit}")

                # Calcul du total
                total = quantity * prix_unitaire if prix_unitaire else 0
                total_global += total

                # Ajout des données au tableau
                all_lots.append({
                    'lot_name': lot_name,
                    'description': description,
                    'quantity': quantity,
                    'unit': unit,
                    'unit_price': prix_unitaire if prix_unitaire else 0,
                    'total': total
                })
            except Exception as e:
                print(f"Erreur lors du traitement du lot {lots[i]} : {e}")
                all_lots.append({
                    'lot_name': lots[i],
                    'description': descriptions[i],
                    'quantity': quantity,
                    'unit': "N/A",
                    'unit_price': 0,
                    'total': 0
                })

        # Construire les données des sous-détails
        all_sub_details = []
        for i in range(len(sub_descriptions)):
            try:
                quantity = float(sub_quantities[i]) if sub_quantities[i] else 0
                unit_price = float(sub_unit_prices[i]) if sub_unit_prices[i] else 0
                total = quantity * unit_price
                total_global += total
                all_sub_details.append({
                    'description': sub_descriptions[i],
                    'quantity': quantity,
                    'unit': sub_units[i],
                    'unit_price': unit_price,
                    'total': total
                })
            except (ValueError, IndexError):
                print(f"Erreur lors du traitement du sous-détail {sub_descriptions[i]}")
                continue

        # Stocker dans la session
        session['all_lots'] = all_lots
        session['all_sub_details'] = all_sub_details
        session['total_global'] = total_global

        return render_template('results.html', lots=all_lots, sub_details=all_sub_details, total_global=total_global)

    return render_template('data_entry.html')

@app.route('/export')
def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Récupération des données depuis la session
    all_lots = session.get('all_lots', [])
    all_sub_details = session.get('all_sub_details', [])
    total_global = session.get('total_global', 0)

    # Création du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "DPGF"

    # Définition des styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="007BFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True, size=12)

    # Titre principal
    ws.merge_cells("A1:F1")
    ws["A1"] = "Décomposition du Prix Global et Forfaitaire (DPGF)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # En-têtes
    headers = ["Lot", "Description", "Quantité", "Unité", "Prix Unitaire (€)", "Total (€)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Insertion des données
    current_row = 4

    # Lots principaux
    for lot in all_lots:
        ws.cell(row=current_row, column=1, value=lot['lot_name'])
        ws.cell(row=current_row, column=2, value=lot['description'])
        ws.cell(row=current_row, column=3, value=lot['quantity'])
        ws.cell(row=current_row, column=4, value=lot['unit'])
        ws.cell(row=current_row, column=5, value=lot['unit_price'])
        ws.cell(row=current_row, column=6, value=lot['total'])

        # Appliquer les styles pour chaque cellule
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = center_align

        current_row += 1

    # Sous-détails
    for sub in all_sub_details:
        ws.cell(row=current_row, column=1, value="Sous-Détail")
        ws.cell(row=current_row, column=2, value=sub['description'])
        ws.cell(row=current_row, column=3, value=sub['quantity'])
        ws.cell(row=current_row, column=4, value=sub['unit'])
        ws.cell(row=current_row, column=5, value=sub['unit_price'])
        ws.cell(row=current_row, column=6, value=sub['total'])

        # Appliquer les styles pour chaque cellule
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = center_align

        current_row += 1

    # Total global
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=1, value="TOTAL GLOBAL").font = bold_font
    ws.cell(row=current_row, column=6, value=total_global).font = bold_font

    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border
        cell.alignment = center_align

    # Ajustement des largeurs de colonnes
    column_widths = [20, 40, 15, 15, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Sauvegarde dans un fichier en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="DPGF_Resultats_Dynamiques.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == '__main__':
    app.run(debug=True)
